from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import BinaryIO

from openpyxl import load_workbook


@dataclass(frozen=True)
class ImportedPersonnel:
    name: str
    extension: str | None


# Uzun / spesifik basliklar once. Kisa "ad" yalnizca tam eslesmede.
_NAME_HEADERS_EXACT = frozenset(
    {
        "ad",
        "adi",
        "isim",
        "name",
        "personel",
        "personelad",
        "personeladi",
        "personelismi",
        "personeladi",
        "agent",
        "agentname",
    }
)
_EXT_HEADERS_EXACT = frozenset(
    {
        "dahili",
        "dahilino",
        "dahiliad",
        "extension",
        "extensionnumber",
        "ext",
        "numara",
        "no",
    }
)
# Icerik eslesmesi icin min 4 karakter (adi/ad false positive olmasin)
_NAME_HEADERS_FUZZY = (
    "personelismi",
    "personeladi",
    "personelad",
    "personel",
    "agentname",
)
_EXT_HEADERS_FUZZY = (
    "extensionnumber",
    "dahilino",
    "dahiliad",
    "extension",
    "dahili",
)


def parse_personnel_workbook(source: bytes | BinaryIO) -> list[ImportedPersonnel]:
    if isinstance(source, bytes):
        source = BytesIO(source)
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook.active
    rows = [tuple(row) for row in sheet.iter_rows(values_only=True)]
    rows = [row for row in rows if any(cell not in (None, "") for cell in row)]
    if not rows:
        return []

    header_mode, name_idx, ext_idx = _detect_layout(rows[0])
    data_rows = rows[1:] if header_mode else rows

    results: list[ImportedPersonnel] = []
    seen_names: set[str] = set()
    for row in data_rows:
        if not row or name_idx >= len(row):
            continue
        name = _clean_name(row[name_idx])
        if not name:
            continue
        # Yanlis kolon: isim yerine sadece sayi geldiyse atla
        if _looks_like_extension_only(name):
            continue
        extension = None
        if ext_idx is not None and ext_idx < len(row) and row[ext_idx] not in (None, ""):
            extension = _clean_extension(row[ext_idx])
        key = name.casefold()
        if key in seen_names:
            continue
        seen_names.add(key)
        results.append(ImportedPersonnel(name=name, extension=extension))
    return results


def _detect_layout(first_row: tuple) -> tuple[bool, int, int | None]:
    """(header_mode, name_idx, ext_idx)."""
    headers = [_normalize_header(cell) for cell in first_row]
    name_idx = _find_column(headers, _NAME_HEADERS_EXACT, _NAME_HEADERS_FUZZY)
    ext_idx = _find_column(headers, _EXT_HEADERS_EXACT, _EXT_HEADERS_FUZZY)

    # Guclu sinyal: 2. kolon sayisal, 1. kolon bilinen baslik DEGIL -> basliksiz personel satiri
    if len(first_row) >= 2 and _looks_like_extension_only(str(first_row[1] or "").strip()):
        if not _is_exact_or_fuzzy_header(headers[0]):
            return False, 0, 1

    # 1. kolon sayisal dahili, 2. kolon isim (personel listesi header'siz)
    if len(first_row) >= 2 and _looks_like_extension_only(str(first_row[0] or "").strip()):
        if not _is_exact_or_fuzzy_header(headers[1] if len(headers) > 1 else ""):
            # data: ext, name - ama header yoksa
            # Eger baslik varsa (DAHILI AD, PERSONEL ISMI) asagidaki header_mode yakalar
            if name_idx is None and ext_idx is None:
                return False, 1, 0

    if name_idx is not None and _is_exact_or_fuzzy_header(headers[name_idx]):
        # Isim ve dahili ayni kolona dustuyse ayir
        if ext_idx is not None and name_idx == ext_idx:
            alt = _find_column(
                headers,
                frozenset({"personelismi", "personeladi", "personel", "isim", "name", "adi", "ad"}),
                ("personelismi", "personeladi", "personel"),
            )
            if alt is not None and alt != ext_idx:
                name_idx = alt
            else:
                # dahili ad vs personel ismi
                for i, h in enumerate(headers):
                    if i != ext_idx and _is_exact_or_fuzzy_header(h) and i != name_idx:
                        if h in _NAME_HEADERS_EXACT or any(f in h for f in _NAME_HEADERS_FUZZY):
                            name_idx = i
                            break
        return True, name_idx, ext_idx

    # Baslik yok
    if len(first_row) >= 2 and _looks_like_extension_only(str(first_row[0] or "").strip()):
        return False, 1, 0
    return False, 0, 1 if len(first_row) >= 2 else None


def _find_column(
    headers: list[str],
    exact: frozenset[str],
    fuzzy: tuple[str, ...],
) -> int | None:
    for index, header in enumerate(headers):
        if header in exact:
            return index
    for candidate in fuzzy:
        for index, header in enumerate(headers):
            if not header:
                continue
            if header == candidate or candidate in header:
                return index
    return None


def _is_exact_or_fuzzy_header(normalized: str) -> bool:
    if not normalized:
        return False
    if normalized in _NAME_HEADERS_EXACT or normalized in _EXT_HEADERS_EXACT:
        return True
    for candidate in _NAME_HEADERS_FUZZY + _EXT_HEADERS_FUZZY:
        if candidate in normalized or normalized == candidate:
            return True
    return False


def _normalize_header(value: object) -> str:
    text = str(value or "").strip().casefold().replace("i̇", "i")
    replacements = str.maketrans(
        {
            "ç": "c",
            "ğ": "g",
            "ı": "i",
            "ö": "o",
            "ş": "s",
            "ü": "u",
            " ": "",
            "_": "",
            "-": "",
        }
    )
    return text.translate(replacements)


def _clean_name(value: object) -> str:
    return " ".join(str(value or "").strip().split())


def _clean_extension(value: object) -> str | None:
    text = str(value).strip()
    if not text or text.lower() in {"none", "null", "#n/a", "nan"}:
        return None
    if text.endswith(".0"):
        text = text[:-2]
    try:
        number = float(text.replace(",", "."))
        if number.is_integer():
            return str(int(number))
    except ValueError:
        pass
    return text or None


def _looks_like_extension_only(value: str) -> bool:
    text = (value or "").strip()
    if not text:
        return False
    try:
        float(text.replace(",", "."))
        return True
    except ValueError:
        return False
